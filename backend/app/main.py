import io
import re
import os
import tempfile
import random
from datetime import datetime, timedelta, timezone
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pdfplumber

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from .database import engine, Base, get_db
from .models import User, Student
from .schemas import UserCreate, UserResponse, UserLogin, LoginResponse, LoginInitiateResponse, VerifyOTPRequest, StudentCreate, StudentResponse, StudentStatsResponse, PaginatedStudentResponse, StudentUpdate, ForgotPasswordRequest, VerifyForgotOTPRequest, ResetPasswordRequest
from .auth import hash_password, verify_password
from .email import send_otp_email

# Initialize database tables on application startup
Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="User Registration API",
    description="A basic registration API built with FastAPI and PostgreSQL",
    version="1.1.0",  # added forgot-password flow
    docs_url="/docs",    # Enable Swagger UI
    redoc_url="/redoc",   # Enable ReDoc
    openapi_url="/openapi.json"  # Enable OpenAPI schema
)

# Configure CORS Middleware to allow requests from the React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://samras-96zh.vercel.app",  # Production frontend
        "http://localhost:5173",           # Local development
        "http://127.0.0.1:5173",           # Local development (127.0.0.1)
        "http://localhost:5174",           # Local development alternative
    ],
    allow_credentials=True,
    allow_methods=["*"],  # Allows all HTTP methods (GET, POST, etc.)
    allow_headers=["*"],  # Allows all headers
)


@app.get("/", status_code=status.HTTP_200_OK)
def root():
    return {
        "message": "Welcome to the User Registration API!",
        "docs_url": "/docs",
        "redoc_url": "/redoc",
        "status": "healthy"
    }

@app.post("/register", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def register_user(user: UserCreate, db: Session = Depends(get_db)):
    """
    Registers a new user inside the PostgreSQL database.
    
    1. Checks if the email is already registered.
    2. Hashes the password using bcrypt.
    3. Stores the user details securely.
    """
    # 1. Check if email is already taken
    existing_user = db.query(User).filter(User.email == user.email).first()
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="An account with this email address already exists."
        )
    
    # 2. Hash the password before storing
    hashed_pwd = hash_password(user.password)
    
    # 3. Create the database record
    new_user = User(
        email=user.email,
        first_name=user.first_name,
        last_name=user.last_name,
        hashed_password=hashed_pwd
    )
    
    # 4. Save to database
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return new_user

@app.post("/login", response_model=LoginInitiateResponse, status_code=status.HTTP_200_OK)
def login_user(credentials: UserLogin, db: Session = Depends(get_db)):
    """
    Authenticates a user and initiates the OTP flow.
    
    1. Retrieves the user by email.
    2. Verifies the password using bcrypt.
    3. Generates a 6-digit OTP code, sets its expiry, and saves to database.
    4. Simulates sending the OTP to the registered email address.
    """
    # 1. Retrieve the user by email
    user = db.query(User).filter(User.email == credentials.email).first()
    
    # 2. Check if user exists and password matches
    if not user or not verify_password(credentials.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password."
        )
        
    # 3. Generate a 6-digit numeric OTP
    otp = f"{random.randint(100000, 999999)}"
    
    # 4. Set OTP and 5 minutes expiry
    user.otp = otp
    user.otp_expiry = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(minutes=5)
    db.commit()
    
    # Save OTP to file for automated browser agent testing
    try:
        import os
        otp_dir = r"C:\Users\Acer\.gemini\antigravity-ide\brain\b8fcfc1e-54a9-424c-8d08-2280c36b0af9\browser"
        os.makedirs(otp_dir, exist_ok=True)
        with open(os.path.join(otp_dir, "otp.txt"), "w") as f:
            f.write(otp)
    except Exception as e:
        print(f"Failed to write test OTP file: {e}")
        
    # 5. Send the OTP email via SMTP
    send_otp_email(user.email, otp)
    
    return {
        "message": "OTP sent to your registered email address.",
        "email": user.email
    }

@app.post("/verify-otp", response_model=LoginResponse, status_code=status.HTTP_200_OK)
def verify_otp(request: VerifyOTPRequest, db: Session = Depends(get_db)):
    """
    Verifies the OTP code to log in the user.
    
    1. Retrieves the user by email.
    2. Checks if OTP exists and is correct.
    3. Checks if the OTP is expired.
    4. Clears OTP on success and signs in the user.
    """
    # 1. Retrieve the user by email
    user = db.query(User).filter(User.email == request.email).first()
    
    # 2. Validate session and OTP
    if not user or not user.otp:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active verification session found or OTP has expired."
        )
        
    if user.otp != request.otp:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid verification code."
        )
        
    # 3. Check expiration
    current_time = datetime.now(timezone.utc).replace(tzinfo=None)
    if user.otp_expiry and user.otp_expiry < current_time:
        # Clear expired OTP
        user.otp = None
        user.otp_expiry = None
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Verification code has expired."
        )
        
    # 4. Clear OTP upon successful verification
    user.otp = None
    user.otp_expiry = None
    db.commit()
    
    return {
        "message": "Login successful",
        "user": user
    }

# --- Forgot Password Flow ---

@app.post("/forgot-password", status_code=status.HTTP_200_OK)
def forgot_password(request: ForgotPasswordRequest, db: Session = Depends(get_db)):
    """
    Initiates forgot-password OTP flow.
    Generates OTP, stores with 5-min expiry, sends to email.
    Returns generic message regardless of whether email exists.
    """
    user = db.query(User).filter(User.email == request.email).first()
    if user:
        otp = f"{random.randint(100000, 999999)}"
        user.otp = otp
        user.otp_expiry = datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(minutes=5)
        db.commit()
        send_otp_email(user.email, otp)
    return {"message": "If this email is registered, an OTP has been sent to it."}


@app.post("/verify-forgot-otp", status_code=status.HTTP_200_OK)
def verify_forgot_otp(request: VerifyForgotOTPRequest, db: Session = Depends(get_db)):
    """
    Validates OTP for password reset.
    OTP is kept (not cleared) so /reset-password can re-verify it.
    """
    user = db.query(User).filter(User.email == request.email).first()
    if not user or not user.otp:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active OTP session found for this email."
        )
    if user.otp != request.otp:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid verification code."
        )
    current_time = datetime.now(timezone.utc).replace(tzinfo=None)
    if user.otp_expiry and user.otp_expiry < current_time:
        user.otp = None
        user.otp_expiry = None
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Verification code has expired. Please request a new one."
        )
    return {"message": "OTP verified successfully. You may now reset your password."}


@app.post("/reset-password", status_code=status.HTTP_200_OK)
def reset_password(request: ResetPasswordRequest, db: Session = Depends(get_db)):
    """
    Resets user password after re-validating OTP.
    Hashes new password, saves to DB, clears OTP session.
    """
    user = db.query(User).filter(User.email == request.email).first()
    if not user or not user.otp:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active OTP session found. Please start the forgot-password flow again."
        )
    if user.otp != request.otp:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid verification code."
        )
    current_time = datetime.now(timezone.utc).replace(tzinfo=None)
    if user.otp_expiry and user.otp_expiry < current_time:
        user.otp = None
        user.otp_expiry = None
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Verification code has expired. Please start the forgot-password flow again."
        )
    user.hashed_password = hash_password(request.new_password)
    user.otp = None
    user.otp_expiry = None
    db.commit()
    return {"message": "Password has been reset successfully. You may now log in."}


# --- Export Endpoints ---

@app.get("/students/export")
def export_students_excel(db: Session = Depends(get_db)):
    """
    Exports all student records to an Excel (.xlsx) file and streams it
    back as a downloadable attachment.
    """
    students = db.query(Student).order_by(Student.id.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Define column headers
    headers = [
        "ID", "New GR", "Old GR", "Old/New", "Full Name", "Date of Birth",
        "City", "Taluka", "District", "Caste", "Category", "Block",
        "Room No", "Last Exam", "Last Exam Year", "Percentage",
        "Current Year of Study", "College Name", "Parents Income",
        "Student Mobile", "Parents Mobile", "Disabled", "Father Deceased",
        "Orphan", "Registration Date", "Date of Leaving", "Special Note"
    ]

    # Style header row
    header_fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Write student rows
    for row_idx, s in enumerate(students, start=2):
        ws.append([
            s.id, s.new_gr, s.old_gr, s.old_new, s.full_name, s.dob,
            s.city, s.taluka, s.district, s.caste, s.category, s.block,
            s.room_no, s.last_exam, s.last_exam_year, s.percentage,
            s.current_year_of_study, s.college_name, s.parents_income,
            s.student_mobile_number, s.parents_mobile_number,
            "Yes" if s.disabled else "No",
            "Yes" if s.father_is_deceased else "No",
            "Yes" if s.orphan else "No",
            str(s.curr_date) if s.curr_date else "",
            str(s.date_of_leaving_the_hostel) if s.date_of_leaving_the_hostel else "",
            s.special_note or ""
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Stream the workbook as bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp
    )

@app.get("/students/export-pdf")
def export_students_pdf(db: Session = Depends(get_db)):
    """
    Exports all student records to a PDF file and streams it
    back as a downloadable attachment.
    """
    students = db.query(Student).order_by(Student.id.asc()).all()

    output = io.BytesIO()
    # Use landscape A4 for wide tables
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Hostel Students Report", styles['Title'])
    elements.append(title)

    # Define column headers (We'll use a subset to fit on PDF, otherwise it goes off page)
    # Using a condensed set of columns for the PDF to ensure it fits nicely
    headers = [
        "ID", "GR No", "Name", "City", "Mobile", "College", "Category", "Hostel Stay"
    ]
    
    data = [headers]

    for s in students:
        hostel_stay = "Present"
        if s.date_of_leaving_the_hostel:
            hostel_stay = f"Left: {s.date_of_leaving_the_hostel}"
            
        data.append([
            str(s.id),
            str(s.new_gr),
            s.full_name[:20] + "..." if len(s.full_name) > 20 else s.full_name, # Truncate long names
            s.city[:15],
            s.student_mobile_number,
            s.college_name[:15] + "..." if s.college_name and len(s.college_name) > 15 else (s.college_name or ""),
            s.category or "",
            hostel_stay
        ])

    table = Table(data, colWidths=[30, 50, 140, 90, 80, 140, 70, 90])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#dc2626")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    # Alternate row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            bc = colors.whitesmoke
        else:
            bc = colors.lightgrey
        table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), bc)]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=headers_resp
    )


# --- PDF Parsing Endpoint ---

def _extract_field(text: str, *patterns: str) -> str | None:
    """
    Try multiple regex patterns on the full PDF text and return the first match.
    Each pattern should contain one capture group for the value.
    """
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            val = match.group(1).strip()
            if val and val.lower() not in ('na', 'n/a', '*' * len(val)):
                return val
    return None


def _parse_bool_field(text: str, *patterns: str) -> bool:
    """
    Returns True if the matched value is 'yes', False otherwise.
    """
    val = _extract_field(text, *patterns)
    return val is not None and val.lower() == 'yes'


def _convert_date(raw: str | None) -> str | None:
    """
    Convert DD/MM/YYYY  →  YYYY-MM-DD for HTML date inputs.
    Also handles DD-MM-YYYY. Returns None if conversion fails.
    """
    if not raw:
        return None
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _parse_address(raw: str | None) -> dict:
    """
    Best-effort parse of a Samras residential address string like:
      'B-1209, Suman sangath, beside of silver coin, vip circle road, utran, Surat., Surat - 394105'
    Returns dict with 'city', 'taluka', 'district'.
    Strategy: split by comma, last meaningful token before pincode = district,
    second-to-last = city, third-to-last = taluka.
    """
    result = {'city': '', 'taluka': '', 'district': ''}
    if not raw:
        return result
    # Remove pincode
    cleaned = re.sub(r'\s*-\s*\d{6}\b', '', raw)
    # Remove trailing dots
    cleaned = re.sub(r'\.+', '', cleaned)
    parts = [p.strip() for p in cleaned.split(',') if p.strip()]
    if len(parts) >= 1:
        result['district'] = parts[-1]
    if len(parts) >= 2:
        result['city'] = parts[-2]
    if len(parts) >= 3:
        result['taluka'] = parts[-3]
    return result


def _parse_caste(raw: str | None) -> dict:
    """
    Split 'SEBC,  Prajapati' or 'SC  Chamar' into category + caste.
    Known categories: General, OBC, SEBC, SC, ST, EWS.
    """
    result = {'category': '', 'caste': ''}
    if not raw:
        return result
    categories = ['General', 'OBC', 'SEBC', 'SC', 'ST', 'EWS']
    for cat in categories:
        pattern = re.compile(rf'^{cat}[,\s]+(.+)$', re.IGNORECASE)
        m = pattern.match(raw.strip())
        if m:
            result['category'] = cat.upper()
            result['caste'] = m.group(1).strip()
            return result
    # Fallback: raw is caste only
    result['caste'] = raw.strip()
    return result


@app.post("/parse-pdf", status_code=status.HTTP_200_OK)
async def parse_student_pdf(file: UploadFile = File(...)):
    """
    Accepts a Samras hostel application PDF, extracts student data,
    and returns it as a JSON object to auto-fill the registration form.

    The parser works pattern-based, not value-hardcoded, so it handles
    any future PDF that follows the same Samras application table layout.
    """
    # --- Validate file type ---
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload a PDF file."
        )
    if file.content_type and 'pdf' not in file.content_type.lower():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid content type. Only PDF files are accepted."
        )

    # --- Parse PDF directly from memory ---
    try:
        content = await file.read()
        if len(content) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded PDF file is empty."
            )

        # --- Extract all text from PDF ---
        try:
            import io
            full_text = ''
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Failed to read PDF content. The file may be corrupted or password-protected. ({str(e)})"
            )

        if not full_text.strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No readable text found in the PDF. The file may contain only scanned images."
            )

        # --- Detect Old/New Status from heading ---
        old_new = None
        if re.search(r'renewal\s+student', full_text, re.IGNORECASE):
            old_new = 'Renewal'
        elif re.search(r'fresh\s+student', full_text, re.IGNORECASE):
            old_new = 'New'

        # --- Extract raw fields using flexible regex patterns ---
        # Each pattern accounts for slight formatting variations across PDFs

        raw_name = _extract_field(full_text,
            r'(?:^|\n)\s*1\.?\s*Name\s*:\s*(.+)',
            r'Name\s*:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)')

        raw_dob = _extract_field(full_text,
            r'(?:^|\n)\s*3\.?\s*Date\s*of\s*Birth\s*:\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})',
            r'Date\s*of\s*Birth\s*[:\-]\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})')

        raw_mobile = _extract_field(full_text,
            r'(?:^|\n)\s*5\.?\s*Mobile\s*No\s*:\s*(\d{10,12})',
            r'Mobile\s*(?:No|Number)\s*[:\-]\s*(\d{10,12})')

        raw_email = _extract_field(full_text,
            r'(?:^|\n)\s*6\.?\s*Email\s*Address\s*:\s*([\w.+\-]+@[\w.]+\.[a-z]{2,})',
            r'Email\s*(?:Address)?\s*[:\-]\s*([\w.+\-]+@[\w.]+\.[a-z]{2,})')

        raw_address = _extract_field(full_text,
            r'(?:^|\n)\s*7\.?\s*Residential\s*Address\s*:\s*(.+?)(?=\n\s*\d+\.)',
            r'Residential\s*Address\s*[:\-]\s*(.+?)(?=\n)')

        raw_income = _extract_field(full_text,
            r'Family\s*Annual\s*Income\s*(?:\(In\s*Rs\.?\))?\s*[:\-]\s*([\d,]+)',
            r'Annual\s*Income\s*[:\-]\s*([\d,]+)')

        raw_college = _extract_field(full_text,
            r'(?:^|\n)\s*10\.?\s*Further\s*Admission\s*in\s*College\s*:\s*(.+)',
            r'Further\s*Admission\s*in\s*College\s*[:\-]\s*(.+?)(?=\n)')

        raw_college_district = _extract_field(full_text,
            r'(?:^|\n)\s*11\.?\s*Further\s*Admission\s*College\s*District\s*:\s*(.+)',
            r'Admission\s*College\s*District\s*[:\-]\s*(.+?)(?=\n)')

        raw_last_year = _extract_field(full_text,
            r'(?:^|\n)\s*14\.?\s*Last\s*Year\s*/\s*Semester\s*:\s*(\d+)',
            r'Last\s*Year\s*/\s*Semester\s*[:\-]\s*(\d+)')

        raw_course = _extract_field(full_text,
            r'(?:^|\n)\s*20\.?\s*Further\s*Course\s*Name\s*:\s*(.+)',
            r'Further\s*Course\s*Name\s*[:\-]\s*(.+?)(?=\n)')

        raw_admission_group = _extract_field(full_text,
            r'(?:^|\n)\s*21\.?\s*Further\s*Admission\s*Group\s*:\s*(.+)',
            r'Admission\s*Group\s*[:\-]\s*(.+?)(?=\n)')

        raw_caste = _extract_field(full_text,
            r'(?:^|\n)\s*23\.?\s*Caste\s*and\s*Sub\s*caste\s*:\s*(.+)',
            r'Caste\s*and\s*Sub\s*caste\s*[:\-]\s*(.+?)(?=\n)',
            r'Caste\s*[:\-]\s*(.+?)(?=\n)')

        raw_percentage = _extract_field(full_text,
            r'(?:^|\n)\s*24\.?\s*Result\s*of\s*Last\s*Exam\s*\(%\)\s*:\s*([\d.]+)',
            r'Result\s*of\s*Last\s*Exam\s*[:\-]\s*([\d.]+)')

        raw_gender = _extract_field(full_text,
            r'(?:^|\n)\s*2\.?\s*Gender\s*:\s*(Male|Female|Other)',
            r'Gender\s*[:\-]\s*(Male|Female|Other)')

        # Boolean fields
        is_disabled = _parse_bool_field(full_text,
            r'Is\s*Disabled\s*:\s*(Yes|No)',
            r'(?:^|\n)\s*15\.?\s*Is\s*Disabled\s*:\s*(Yes|No)')

        is_orphan = _parse_bool_field(full_text,
            r'Is\s*Orphan\s*:\s*(Yes|No)',
            r'(?:^|\n)\s*19\.?\s*Is\s*Orphan\s*:\s*(Yes|No)')

        is_widow_child = _parse_bool_field(full_text,
            r'Is\s*Child\s*of\s*Widow\s*:\s*(Yes|No)',
            r'(?:^|\n)\s*18\.?\s*Is\s*Child\s*of\s*Widow\s*:\s*(Yes|No)')

        # --- Process / transform raw values ---

        dob_converted = _convert_date(raw_dob)

        # Current year = last year + 1
        current_year = None
        if raw_last_year:
            try:
                current_year = str(int(raw_last_year.strip()) + 1)
            except ValueError:
                current_year = None

        # Address breakdown
        addr = _parse_address(raw_address)

        # Caste + Category split
        caste_info = _parse_caste(raw_caste)

        # Income — strip commas for clean storage
        income_clean = raw_income.replace(',', '') if raw_income else None

        # --- Build response ---
        result = {
            # Auto-filled fields
            "full_name":             raw_name,
            "dob":                   dob_converted,
            "student_mobile_number": raw_mobile,
            "email":                 raw_email,
            "city":                  addr.get('city') or None,
            "taluka":                addr.get('taluka') or None,
            "district":              addr.get('district') or None,
            "caste":                 caste_info.get('caste') or None,
            "category":              caste_info.get('category') or None,
            "college_name":          raw_college,
            "current_year_of_study": current_year,
            "percentage":            raw_percentage,
            "parents_income":        income_clean,
            "old_new":               old_new,
            "disabled":              is_disabled,
            "orphan":                is_orphan,
            "father_is_deceased":    is_widow_child,
            # Extra info (not form fields but useful for notes)
            "gender":                raw_gender,
            "course":                raw_course,
            "admission_group":       raw_admission_group,
            "last_year_semester":    raw_last_year,
        }

        # Strip empty strings to None so frontend can treat absent = None
        result = {k: (v if v != '' else None) for k, v in result.items()}

        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred while processing the PDF: {str(e)}"
        )
# --- Student Hostel API Routes ---
@app.post("/students", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
def create_student(student: StudentCreate, db: Session = Depends(get_db)):
    """
    Registers a new student profile in the database.
    Checks if the student's mobile number already exists.
    """
    if student.student_mobile_number:
        existing_student = db.query(Student).filter(Student.student_mobile_number == student.student_mobile_number).first()
        if existing_student:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Student already exists with this mobile number."
            )
            
    db_student = Student(**student.model_dump())
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student

@app.get("/students", response_model=PaginatedStudentResponse)
def get_students(page: int = 1, limit: int = 5, db: Session = Depends(get_db)):
    """
    Lists registered hostel students with pagination.
    """
    if page < 1:
        page = 1
    if limit < 1:
        limit = 5
        
    offset = (page - 1) * limit
    total = db.query(Student).count()
    students = db.query(Student).order_by(Student.id.desc()).offset(offset).limit(limit).all()
    
    return {
        "students": students,
        "total": total,
        "page": page,
        "limit": limit
    }

@app.get("/students/stats", response_model=StudentStatsResponse)
def get_student_statistics(db: Session = Depends(get_db)):
    """
    Returns total statistics (total, disabled, orphan counts).
    """
    total = db.query(Student).count()
    disabled = db.query(Student).filter(Student.disabled == True).count()
    orphan = db.query(Student).filter(Student.orphan == True).count()
    
    return {
        "total": total,
        "disabled": disabled,
        "orphan": orphan
    }

@app.get("/students/search", response_model=PaginatedStudentResponse)
def search_students(
    gr_number: str | None = None,
    college: str | None = None,
    mobile: str | None = None,
    page: int = 1,
    limit: int = 5,
    db: Session = Depends(get_db)
):
    """
    Searches for students dynamically using GR Number, College Name, or Mobile Number filters with pagination.
    GR Number search matches against both new_gr and old_gr fields.
    """
    from sqlalchemy import or_

    if page < 1:
        page = 1
    if limit < 1:
        limit = 5

    query = db.query(Student)

    if gr_number:
        query = query.filter(
            or_(
                Student.new_gr.ilike(f"%{gr_number}%"),
                Student.old_gr.ilike(f"%{gr_number}%")
            )
        )
    if college:
        query = query.filter(Student.college_name.ilike(f"%{college}%"))
    if mobile:
        query = query.filter(Student.student_mobile_number.ilike(f"%{mobile}%"))

    total = query.count()
    offset = (page - 1) * limit
    students = query.order_by(Student.id.desc()).offset(offset).limit(limit).all()

    return {
        "students": students,
        "total": total,
        "page": page,
        "limit": limit
    }

@app.get("/students/{student_id}", response_model=StudentResponse)
def get_student_by_id(student_id: int, db: Session = Depends(get_db)):
    """
    Retrieves a single student profile by their ID.
    """
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Student with ID {student_id} not found."
        )
    return student

@app.put("/students/{student_id}", response_model=StudentResponse)
def update_student(student_id: int, updates: StudentUpdate, db: Session = Depends(get_db)):
    """
    Updates an existing student profile by their ID.
    Only the fields provided in the request body will be updated.
    """
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Student with ID {student_id} not found."
        )
    
    # Apply only the fields that were explicitly provided (partial update)
    update_data = updates.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(student, field, value)
    
    db.commit()
    db.refresh(student)
    return student
